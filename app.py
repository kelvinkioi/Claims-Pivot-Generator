import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="Pivot Table Generator", layout="centered")
st.title("📊 Pivot Table Generator")

# ------------------------------
# Step 1: Upload and Cache the File
# ------------------------------
uploaded_file = st.file_uploader(
    "Upload your Excel file (must contain columns: SCHEME, TRANSACTION DATE, BENEFIT, AMOUNT, COUNT, UNIQUE COUNT, PROVIDER NAME)",
    type=["xlsx"]
)

if uploaded_file:
    if "df" not in st.session_state:
        with st.spinner("Reading file..."):
            try:
                st.session_state.df = pd.read_excel(uploaded_file, sheet_name="Sheet1")
                st.success("✅ File loaded successfully!")
            except Exception as e:
                st.error(f"Failed to read the Excel file: {e}")
                st.stop()
    
    df = st.session_state.df
    required_cols = ['SCHEME', 'TRANSACTION DATE', 'BENEFIT', 'AMOUNT', 'COUNT', 'UNIQUE COUNT', 'PROVIDER NAME']
    if not all(col in df.columns for col in required_cols):
        st.error(f"Missing required columns. Expected columns: {required_cols}")
        st.stop()

    df['TRANSACTION DATE'] = pd.to_datetime(df['TRANSACTION DATE'], errors='coerce')
    unique_schemes = sorted(df['SCHEME'].dropna().unique().tolist())

    with st.form("input_form"):
        st.info("Select the schemes and set the date range for each (or choose to ignore dates).")
        selected_schemes = st.multiselect("Search and select schemes:", options=unique_schemes)
        
        scheme_date_filters = {}
        if selected_schemes:
            st.markdown("### Configure Date Filters for Each Selected Scheme")
            for scheme in selected_schemes:
                with st.expander(f"Date settings for: **{scheme}**", expanded=True):
                    ignore_dates = st.checkbox(f"Process {scheme} without date filter", key=f"{scheme}_ignore")
                    if not ignore_dates:
                        start_date = st.date_input(f"Start Date for {scheme}", key=f"{scheme}_start", value=datetime.today())
                        end_date = st.date_input(f"End Date for {scheme}", key=f"{scheme}_end", value=datetime.today())
                        scheme_date_filters[scheme] = {
                            "apply_dates": True,
                            "start_date": pd.to_datetime(start_date),
                            "end_date": pd.to_datetime(end_date)
                        }
                    else:
                        scheme_date_filters[scheme] = {"apply_dates": False, "start_date": None, "end_date": None}
        else:
            st.warning("Please select at least one scheme.")

        submitted = st.form_submit_button("Generate Pivot Tables")

    if submitted:
        for scheme, settings in scheme_date_filters.items():
            if settings["apply_dates"] and (settings["start_date"] > settings["end_date"]):
                st.error(f"For scheme {scheme}: Start Date must be before End Date.")
                st.stop()

        with st.spinner("Generating pivot tables. This might take a moment..."):
            output = BytesIO()
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

            for scheme in selected_schemes:
                scheme_df = df[df['SCHEME'] == scheme].copy()
                settings = scheme_date_filters.get(scheme, {"apply_dates": False})
                if settings["apply_dates"]:
                    scheme_df = scheme_df[
                        (scheme_df['TRANSACTION DATE'] >= settings["start_date"]) &
                        (scheme_df['TRANSACTION DATE'] <= settings["end_date"])
                    ]
                if scheme_df.empty:
                    st.warning(f"No data for scheme '{scheme}' in the selected date range. Skipping.")
                    continue

                scheme_df = scheme_df.sort_values(by='TRANSACTION DATE')
                scheme_df['TRANSACTION DATE NORMALIZED'] = scheme_df['TRANSACTION DATE'].dt.strftime('%m/%Y')

                # Pivot 1: Benefit by Amount
                pivot1 = pd.pivot_table(
                    scheme_df, values='AMOUNT', index='TRANSACTION DATE NORMALIZED',
                    columns='BENEFIT', aggfunc='sum', margins=True, margins_name='Grand Total'
                )

                # Pivot 2: Benefit by Count
                pivot2 = pd.pivot_table(
                    scheme_df, values='COUNT', index='TRANSACTION DATE NORMALIZED',
                    columns='BENEFIT', aggfunc='sum', margins=True, margins_name='Grand Total'
                )

                # Pivot 3: Unique Count
                pivot3 = pd.pivot_table(
                    scheme_df, values='UNIQUE COUNT', index='TRANSACTION DATE NORMALIZED',
                    aggfunc='sum', margins=True, margins_name='Grand Total'
                )

                # Pivot 4: Provider Name by Amount (sorted)
                pivot4 = scheme_df.groupby('PROVIDER NAME')['AMOUNT'].sum().sort_values(ascending=False).reset_index()

                # Pivot 5: Provider Name by Count (sorted)
                pivot5 = scheme_df.groupby('PROVIDER NAME')['COUNT'].sum().sort_values(ascending=False).reset_index()

                # Create Excel sheet
                sheet_name = scheme[:31]
                if sheet_name in workbook.sheetnames:
                    i = 1
                    while f"{sheet_name} {i}" in workbook.sheetnames:
                        i += 1
                    sheet_name = f"{sheet_name} {i}"
                sheet = workbook.create_sheet(sheet_name)

                # Dynamic headers for pivot 1 & 2
                dynamic_benefits = sorted(scheme_df['BENEFIT'].unique())
                headers = ['TRANSACTION DATE NORMALIZED'] + dynamic_benefits + ['Grand Total']

                # Write Pivot 1
                sheet.cell(row=1, column=1, value="Benefit by Amount")
                for i, h in enumerate(headers, start=1):
                    sheet.cell(row=2, column=i, value=h)
                for i, (idx, row) in enumerate(pivot1.iterrows(), start=3):
                    sheet.cell(row=i, column=1, value=idx)
                    for j, h in enumerate(headers[1:], start=2):
                        sheet.cell(row=i, column=j, value=row.get(h, 0))

                # Write Pivot 2
                row2_start = pivot1.shape[0] + 5
                sheet.cell(row=row2_start, column=1, value="Benefit by Count")
                for i, h in enumerate(headers, start=1):
                    sheet.cell(row=row2_start + 1, column=i, value=h)
                for i, (idx, row) in enumerate(pivot2.iterrows(), start=row2_start + 2):
                    sheet.cell(row=i, column=1, value=idx)
                    for j, h in enumerate(headers[1:], start=2):
                        sheet.cell(row=i, column=j, value=row.get(h, 0))

                # Write Pivot 3
                row3_start = row2_start + pivot2.shape[0] + 5
                sheet.cell(row=row3_start, column=1, value="Number of Lives (Unique Count)")
                sheet.cell(row=row3_start + 1, column=1, value="TRANSACTION DATE NORMALIZED")
                sheet.cell(row=row3_start + 1, column=2, value="UNIQUE COUNT")
                for i, row in enumerate(pivot3.iterrows(), start=row3_start + 2):
                    sheet.cell(row=i, column=1, value=row[0])
                    sheet.cell(row=i, column=2, value=row[1]['UNIQUE COUNT'])

                # Write Pivot 4
                row4_start = row3_start + pivot3.shape[0] + 5
                sheet.cell(row=row4_start, column=1, value="Provider by Amount (Descending)")
                sheet.cell(row=row4_start + 1, column=1, value="PROVIDER NAME")
                sheet.cell(row=row4_start + 1, column=2, value="AMOUNT")
                for i, row in enumerate(pivot4.itertuples(index=False), start=row4_start + 2):
                    sheet.cell(row=i, column=1, value=row[0])
                    sheet.cell(row=i, column=2, value=row[1])

                # Write Pivot 5
                row5_start = row4_start + pivot4.shape[0] + 5
                sheet.cell(row=row5_start, column=1, value="Provider by Count (Descending)")
                sheet.cell(row=row5_start + 1, column=1, value="PROVIDER NAME")
                sheet.cell(row=row5_start + 1, column=2, value="COUNT")
                for i, row in enumerate(pivot5.itertuples(index=False), start=row5_start + 2):
                    sheet.cell(row=i, column=1, value=row[0])
                    sheet.cell(row=i, column=2, value=row[1])

            workbook.save(output)
            output.seek(0)

        st.success("✅ Pivot tables generated successfully!")
        st.download_button(
            label="📥 Download Pivot Excel File",
            data=output.getvalue(),
            file_name="Pivot_Tables_By_Scheme.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown(
    """
    <hr style='margin-top: 50px; margin-bottom: 10px;'>
    <div style='text-align: center; font-family: Arial, sans-serif;'>
        <p style='margin: 5px; font-size: 16px;'>
            <img src='https://img.shields.io/badge/version-1.08-blue' style='vertical-align: middle; margin-right: 10px;'/>
            <img src='https://img.shields.io/badge/developed-2025-green' style='vertical-align: middle;'/>
        </p>
        <p style='margin: 5px; font-size: 15px;'>
            Developed with ❤️ by 
            <a href='https://github.com/kelvinkioi/' target='_blank' style='text-decoration: none; color: #3366cc;'>
                Kelvin Kioi
            </a>
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

